import sqlite3
import datetime
import openpyxl as op
import csv
import time

with open('202306.csv', 'r', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile)
    for row in reader:

        Money_v = row[0]
        Compa_v = row[3]
        Posts_v = row[4]
        Namel_v = row[1]
        Namef_v = row[2]
        Addrl_v = ''
        conn = sqlite3.connect("meibodata.db")
        cur = conn.cursor()
        cur.execute('create table IF NOT EXISTS omimai(id text,Company_Name text,Post_Name text,Last_Name text,Fast_Name text,Address text,receipt_Money integer,primary key(id))')
        # データの挿入
        dt = datetime.datetime.now()
        dt_now = dt.strftime('%Y%m%d%H%M%S')
        data = [(dt_now,Compa_v,Posts_v,Namel_v,Namef_v,Addrl_v,Money_v)]
        cur.executemany('INSERT INTO omimai VALUES(?,?,?,?,?,?,?)',data)
        cur.close()
        conn.commit()
        conn.close()

        wb = op.load_workbook('omimai.xlsx')
        ws = wb['Sheet1']

        max_row = ws.max_row +1
        ws.cell(row = max_row,column = 1).value = dt_now
        ws.cell(row = max_row,column = 2).value = Namel_v
        ws.cell(row = max_row,column = 3).value = Namef_v
        ws.cell(row = max_row,column = 4).value = Addrl_v
        ws.cell(row = max_row,column = 5).value = Money_v
        ws.cell(row = max_row,column = 6).value = Compa_v
        ws.cell(row = max_row,column = 7).value = Posts_v
        wb.save('omimai.xlsx')
        wb.close()
        time.sleep(2)
