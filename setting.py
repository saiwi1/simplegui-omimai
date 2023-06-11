import openpyxl as op

wb = op.load_workbook('omimai.xlsx')
ws = wb['Sheet1']

max_row = ws.max_row
ws.cell(row = max_row,column = 1).value = 'ID'
ws.cell(row = max_row,column = 2).value = '苗字'
ws.cell(row = max_row,column = 3).value = '名前'
ws.cell(row = max_row,column = 4).value = '住所'
ws.cell(row = max_row,column = 5).value = '金額（千円）'
ws.cell(row = max_row,column = 6).value = '法人名'
ws.cell(row = max_row,column = 7).value = '役職'
wb.save('omimai.xlsx')


