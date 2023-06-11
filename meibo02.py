import PySimpleGUI as sg
import sqlite3
import datetime
import openpyxl as op




sg.theme('BluePurple')

layout = [
          [[sg.Text('金額(1,000)'),sg.Input(key='-Money-',default_text='0', size=(6,1))],
          [sg.Text('姓'),sg.Input(key='-Lname-'),sg.Text('名'),sg.Input(key='-Fname-')],
          [[sg.Text('住所'),sg.Input(key='-Address-')],
          [sg.Text('法人名'),sg.Input(key='-Company-'),sg.Text('役職'),sg.Input(key='-Post-')],
          [sg.Button('登録'), sg.Button('Exit')]]
          ]]
window = sg.Window('Pattern 2B', layout,font='any 18')

while True:  # Event Loop
    event, values = window.read()
    print(event, values)
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    if event == '登録':
        # Update the "output" text element to be the value of "input" element
        #window['-OUTPUT-'].update(values['-IN-'])
        Money_v = int(values['-Money-'])
        Compa_v = values['-Company-']
        Posts_v = values['-Post-']
        Namel_v = values['-Lname-']
        Namef_v = values['-Fname-']
        Addrl_v = values['-Address-']
        conn = sqlite3.connect("meibodata.db")
        cur = conn.cursor()
        cur.execute('create table IF NOT EXISTS omimai(id text,Company_Name text,Post_Name text,Last_Name text,Fast_Name text,Address text,receipt_Money integer,primary key(id))')
        # データの挿入
        dt = datetime.datetime.now()
        dt_now = dt.strftime('%Y%m%d%H%M%S')
        data = [(dt_now,Compa_v,Posts_v,Namel_v,Namef_v,Addrl_v,Money_v)]
        cur.executemany('INSERT INTO omimai VALUES(?,?,?,?,?,?,?)',data)
        cur.close()
        conn.commit()
        conn.close()

        wb = op.load_workbook('omimai.xlsx')
        ws = wb['Sheet1']

        max_row = ws.max_row +1
        ws.cell(row = max_row,column = 1).value = dt_now
        ws.cell(row = max_row,column = 2).value = Namel_v
        ws.cell(row = max_row,column = 3).value = Namef_v
        ws.cell(row = max_row,column = 4).value = Addrl_v
        ws.cell(row = max_row,column = 5).value = Money_v
        ws.cell(row = max_row,column = 6).value = Compa_v
        ws.cell(row = max_row,column = 7).value = Posts_v
        wb.save('omimai.xlsx')
        window['-Money-']. Update('0')
        window['-Company-']. Update('')
        window['-Post-']. Update('')
        window['-Lname-']. Update('')
        window['-Fname-']. Update('')
        window['-Address-']. Update('')



window.close()